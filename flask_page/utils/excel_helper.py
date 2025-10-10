#apps/utils/excel_helper.py
from openpyxl import load_workbook

def af_getexceldict(file_path="static/Nilai Shop Floor Certification Summary.xlsx", tabname="Sheet1"):
    wb = load_workbook(filename=file_path)
    if tabname == "":
        ws = wb.active
    else:
        ws = wb[tabname]
    
    data = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: value for i, value in enumerate(row)}
        data.append(row_data)
    
    return data

def af_addexcel(file_path="tanamxlsx.xlsx", new_row=["ayam","ikan"]):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    ws.append(new_row)
    
    wb.save(file_path)  # Save the workbook after appending the row
    wb.close()