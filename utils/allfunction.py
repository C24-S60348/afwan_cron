import csv
import os
import random

from flask import flash, get_flashed_messages, redirect, render_template_string, request, send_from_directory, url_for
from openpyxl import load_workbook
import pandas as pd
from copy import copy

def af_getexceldict(file_path="static/Nilai Shop Floor Certification Summary.xlsx", tabname="Sheet1"):
    wb = load_workbook(filename=file_path)
    if tabname == "":
        ws = wb.active
    else:
        ws = wb[tabname]
    
    data = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: value for i, value in enumerate(row)}
        data.append(row_data)
    
    return data

def af_addexcel(file_path="tanamxlsx.xlsx", new_row=["ayam","ikan"]):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    ws.append(new_row)
    
    wb.save(file_path)  # Save the workbook after appending the row
    wb.close()

def af_addexcel2(file_path="tanamxlsx.xlsx", data=["ayam","ikan"], desired_row=None):
    # Load the workbook and select the active worksheet
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Create a DataFrame from the provided data list
    new_row = pd.DataFrame([data])

    # Determine the starting row: either specified or append at the end
    if desired_row is not None:
        start_row = desired_row
    else:
        if ws.max_row > 1:
            start_row = ws.max_row + 1
        else:
            start_row = 2
    
    # Iterate over each item in the DataFrame and write it to the sheet
    for idx, row in new_row.iterrows():
        for col_idx, (column_name, value) in enumerate(row.items(), start=1):
            ws.cell(row=start_row + idx, column=col_idx, value=value)

    # Save changes to the workbook and close it
    wb.save(file_path)
    wb.close()

#tak jadi
def af_addexcel3(file_path="tanamxlsx.xlsx", data=["ayam","ikan"], desired_row=None):
    # Load the workbook and select the active worksheet
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Create a DataFrame from the provided data list
    new_row = pd.DataFrame([data])

    # Determine the starting row: either specified or append at the end
    if desired_row is not None:
        start_row = desired_row
    else:
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2

    # Iterate over each item in the DataFrame and write it to the sheet
    for idx, row in new_row.iterrows():
        for col_idx, (column_name, value) in enumerate(row.items(), start=1):
            # Calculate the target cell position
            target_cell = ws.cell(row=start_row + idx, column=col_idx)

            # Use copy to prevent 'unhashable type' error and preserve styles
            original_fill = copy(target_cell.fill)
            original_font = copy(target_cell.font)
            original_border = copy(target_cell.border)
            original_alignment = copy(target_cell.alignment)
            original_number_format = target_cell.number_format

            # Write the new value
            target_cell.value = value

            # Reapply the copied styles to ensure they are not lost
            target_cell.fill = original_fill
            target_cell.font = original_font
            target_cell.border = original_border
            target_cell.alignment = original_alignment
            target_cell.number_format = original_number_format

    # Save changes to the workbook
    wb.save(file_path)
    wb.close()

def af_getcsv(csvloc="tanam.csv"):
    result = []
    with open(csvloc, newline='', encoding='utf-8-sig') as csvfile:
        csvreader = csv.reader(csvfile)
        for row in csvreader:
            result.append(row)
    return result
    # result = af_getcsv("tanam.csv")
    #
    # for r in result:
    #     html += "<li>" + r[0] + "</li>"

def af_getcsvdict(csvloc="tanam.csv"):
    result = []
    with open(csvloc, newline='', encoding='utf-8-sig') as csvfile:
        csvreader = csv.DictReader(csvfile)
        for row in csvreader:
            result.append(row)
    return result
    # result = af_getcsvdict("tanam.csv")
    #
    # for r in result:
    #     html += "<li>" + r["id"] + "</li>"

def af_addcsv(file_path="tanam.csv", new_row=["ikan","ayam"]):
    # Open the CSV file in append mode
    with open(file_path, mode='a', newline='') as file:
        writer = csv.writer(file)
        
        # Write the new row to the file
        writer.writerow(new_row)
    # file_path = 'tanam.csv'
    # new_row = [
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     af_requestpost("tanamtext"),
    #     ]

def af_allowed_file2(filename="flangelist2003.mdb", allowedextensions={'mdb'}):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowedextensions

def af_ifpathnotexist(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)

def af_downloadfile(folder="static", name="flangelist2003.mdb"):
    af_ifpathnotexist(folder)
    try:
        return send_from_directory(folder, name, as_attachment=True)
    except FileNotFoundError:
        return 'File not found', 404
    #return af_downloadfile(UPLOAD_FOLDER, 'flangelist2003.mdb')
    
def af_uploadfile(redirectlink="flangehandle.flangehandle", folder="static", name="flangelist2003.mdb", allowedextensions={'mdb'}):
    
    af_ifpathnotexist(folder)

    if 'file' not in request.files:
        return 'No file part'
    
    file = request.files['file'] #<...name="file">
    
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for(redirectlink))
    
    if file and af_allowed_file2(file.filename, allowedextensions):
        filename = name  # Enforce the filename to be "flange.mdb"
        file.save(os.path.join(folder, filename))
        flash('File successfully uploaded')
        return redirect(url_for(redirectlink))
    
    return 'File not allowed', 400
    #return af_uploadfile("flangehandle.flangehandle", UPLOAD_FOLDER, "flangelist2003.mdb", {'mdb'})

def af_htmlcard(cards=None):
    if cards == None:
        cards = af_htmlcards()
    html = """
    <div class="d-flex align-center justify-content-center my-4">
        """+cards+"""
    </div>
    """
    return html
    # html += af_htmlcard(cards)

def af_htmlcards(href="/pms2", imgsrc="/static/images/pms.png", text="Manufacturing Execution System (MES) Design 2"):
    html = f"""
        <div class="card mx-1">
            <a href="{href}" style="text-decoration: none; color: inherit;">
                <img src="{imgsrc}" class="card-img-top" alt="{text}">
                <div class="card-body">
                    <p class="card-text">{text}</p>
                </div>
            </a>
        </div>
    """
    return html
    # cards += af_htmlcards("/tanam", "/static/images/pms.png", "Tanam")

def af_javascriptchangedropdown(id="style", link="/eqi/qa"):
    html = """
    document.getElementById('"""+id+"""').addEventListener('change', function() {
        // Get selected value
        var styleValue = this.value;
        // Redirect to the desired URL
        window.location.href = '"""+link+"""?"""+id+"""=' + styleValue;
    });
    """
    return html    

def af_htmlcardsuserside50(href="/pms2", imgsrc="/static/images/pms.png", text="Manufacturing Execution System (MES) Design 2"):
    html = f"""
        <div class="card mx-1 w-50">
            <a href="{href}" style="text-decoration: none; color: inherit;display:flex;">
                <img src="{imgsrc}" style="height:140px;width:auto;border-bottom-left-radius: 4px;" class="card-img-top" alt="{text}">
                <div class="card-body">
                    <p class="card-text">{text}</p>
                </div>
            </a>
        </div>
    """
    return html

def af_htmlcardsuserside50center(href="/pms2", imgsrc="/static/images/pms.png", text="Manufacturing Execution System (MES) Design 2"):
    html = f"""
        <div class="card mx-1 w-50">
            <a href="{href}" style="text-decoration: none; color: inherit;text-align:center;">
                <img src="{imgsrc}" style="height:140px;width:auto;border-radius: 4px;" class="card-img-top" alt="{text}">
                <div class="card-body">
                    <p class="card-text">{text}</p>
                </div>
            </a>
        </div>
    """
    return html

def af_htmlcardsuserside50centerfixed(href="/pms2", imgsrc="/static/images/pms.png", text="Manufacturing Execution System (MES) Design 2", width="200px"):
    html = f"""
        <div class="card mx-1" style="width:{width};">
            <a href="{href}" style="text-decoration: none; color: inherit;text-align:center;">
                <img src="{imgsrc}" style="height:140px;width:auto;border-radius: 4px;" class="card-img-top" alt="{text}">
                <div class="card-body">
                    <p class="card-text">{text}</p>
                </div>
            </a>
        </div>
    """
    return html

def af_htmltitle(text="Title"):
    html = f'<div class="title"><h1 class="text-center py-2 my-4">{text}</h1></div>'
    return html
    # html += af_htmltitle("Title")

def af_requestget(code="ikan"):
    if request.args.get(code) == None:
        return ""
    else:
        return request.args.get(code)
    
def af_requestpost(code="ikan", default=""):
    return request.form.get(code, default)

def af_requestpostfromjson(code="ikan", defualt=""):
    rp = request.json
    return rp.get(code, "")

def af_htmlselectoption(value="1", name="First option", selected=""):
    selectedDIV = ""
    if selected != "":
        selectedDIV = "selected"
    html = f"""
        <option value="{value}" {selectedDIV}>{name}</option>
    """
    return html

def af_htmlselectoptionempty():
    html = "<option></option>"
    return html

def af_htmlselect(id="mat", name="Material", options=af_htmlselectoptionempty()):
    html = f"""
    <div>
        <div>{name}:</div>
        <select id="{id}" name="{id}">
           {options}
        </select>
    </div>"""
    return html

def af_htmltextinput(name="ikan", id="ik", placeholder="Put ikan here"):
    html = f"""
        <div>
            <div>{name}</div>
            <input type="text" class="form-control" id="{id}" name="{id}" placeholder="{placeholder}">
        </div>"""
    return html

def af_htmlbutton(name="", typecolor="primary", onclick=""):
    classbtn = ""
    if name == "Back":
        onclick = "window.history.back(); return false;"
    
    if typecolor == "primary":
        classbtn = "btn-primary"
    else:
        classbtn = "btn-secondary"
        
    html = f"""
        <button class="btn {classbtn}" onclick="{onclick}">
            {name}
        </button>
    """
    return html

def af_htmlbuttonlink(name="", typecolor="primary", href=""):
    classbtn = ""
    if typecolor == "primary":
        classbtn = "btn-primary"
    else:
        classbtn = "btn-secondary"
        
    html = f"""
        <a href="{href}" class="btn {classbtn}">
            {name}
        </a>
    """
    return html

def af_htmlformsubmitbutton():
    html = """<button type="submit" class="btn btn-primary">Submit</button>"""
    return html

def af_htmlformstart(link="tanam_add/submit"):
    html = f"<form action='{link}' method='POST'>"
    return html

def af_htmlformend():
    html = "</form>"
    return html

def af_htmlformend2():
    html = ""
    html += af_htmlformsubmitbutton()
    html += "</form>"
    return html

def af_redirect(link="flangehandle.flangehandle"):
    return redirect(url_for(link))

def af_htmlhandletemplate(linkdownload="/flangehandle/download", linkupload="/flangehandle/upload", linkimages="/flangehandle/images", name="flange.mdb", linkback="/flangehandle"):
    html = ""
    html += """
        <!doctype html>
        <html>

            <head>
                <title>Upload/Download File</title>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <link href="{{ url_for('static', filename='bootstrap.min.css') }}" rel="stylesheet" integrity="sha384-sRIl4kxILFvY47J16cr9ZwB07vP4J8+LH7qKQnuqkuIAvNWLzeN8tE5YBujZqJLB" crossorigin="anonymous">
                <style>
                    body {
                        background-color: aliceblue;
                        background-image: url("{{ url_for('static', filename='images/flange.png') }}");
                        background-repeat: no-repeat;
                    }
                    .selectorDIV {
                        max-width: 800px;
                        margin: auto;
                        background-color: rgb(249, 252, 255);
                        margin-top: 20px;
                        padding: 20px;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                        text-align: center;
                    }
                    .chooseDIV div {
                        background-color: #fff;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                        width: 50%;
                        text-align: center;
                        display: flex;
                        flex-direction: column;
                        justify-content: center;
                        align-items: center;
                        padding:10px;
                        margin:5px;
                    }
                    .chooseDIV div form input {
                        margin-bottom: 10px;
                    }
                </style>
            </head>

            <body>
                <div class="selectorDIV" style="
                        font-size: 40px;
                        padding: 5px;
                        margin: 10px;
                        margin-left: 20px;
                        position: absolute;
                    ">
                    <a href=\""""+linkback+"""\" style="
                        text-decoration: none;
                    " >🔙</a>
                </div>
                <div class="container selectorDIV">
                    <div class="d-flex justify-content-center chooseDIV">
                        <div>
                            <h1>📩Download File</h1>
                            <div class="form-control">
                                <a href=\""""+linkdownload+"""\">Download """+name+"""</a>
                            </div>
                        </div>
                        <div>
                            <h1>⬆️Upload File</h1>
                            <form action=\""""+linkupload+"""\" method="post" enctype="multipart/form-data">
                                <input type="file" name="file" class="form-control">
                                <input type="submit" value="Upload" class="btn btn-warning">
                            </form>
                        </div>
                    </div>"""

    html += af_render_flashed_messages()
    
    if linkimages != "":
        html += """
                <div class="d-flex justify-content-center align-items-center my-4">
                    <a href=\""""+linkimages+"""\">
                        <div class="form-control text-center">
                            Images
                        </div>
                    </a>
                </div>
                """
    html += """
            </body>

        </html>
    """
    return render_template_string(html)

def af_render_flashed_messages():
    # Assume get_flashed_messages is defined and returns a list of messages
    messages = get_flashed_messages()
    if not messages:
        return ""

    # Start building the HTML string
    html = "<div class='flashed-messages'><br/><ul>"
    
    # Add each message as a list item
    for message in messages:
        html += f"<li>{message}</li>"
    
    # Close the unordered list and the div
    html += "</ul></div>"
    
    return html


def af_htmlbulatcolor(color="#fabcde"):
    html = f"<div style='border-radius:24px;width:30px;height:30px;background-color:{color};'></div> "
    return html

def af_image_exists(image_path="static/images/employees/afwan.png"):
    # Check if the file exists and is a file (not a directory)
    return os.path.isfile(image_path)


def af_htmlselect(id="mat", name="Material", options=af_htmlselectoptionempty()):
    html = f"""
    <div>
        <div>{name}:</div>
        <select id="{id}" name="{id}">
           {options}
        </select>
    </div>"""
    return html

def af_htmltextinput(name="ikan", id="ik", placeholder="Put ikan here"):
    html = f"""
        <div>
            <div>{name}</div>
            <input type="text" class="form-control" id="{id}" name="{id}" placeholder="{placeholder}">
        </div>"""
    return html

def af_htmlbutton(name="", typecolor="primary", onclick=""):
    classbtn = ""
    if name == "Back":
        onclick = "window.history.back(); return false;"
    
    if typecolor == "primary":
        classbtn = "btn-primary"
    else:
        classbtn = "btn-secondary"
        
    html = f"""
        <button class="btn {classbtn}" onclick="{onclick}">
            {name}
        </button>
    """
    return html

def af_htmlbuttonlink(name="", typecolor="primary", href=""):
    classbtn = ""
    if typecolor == "primary":
        classbtn = "btn-primary"
    else:
        classbtn = "btn-secondary"
        
    html = f"""
        <a href="{href}" class="btn {classbtn}">
            {name}
        </a>
    """
    return html

def af_htmlformsubmitbutton():
    html = """<button type="submit" class="btn btn-primary">Submit</button>"""
    return html

def af_htmlformstart(link="tanam_add/submit"):
    html = f"<form action='{link}' method='POST'>"
    return html

def af_htmlformend():
    html = "</form>"
    return html

def af_htmlformend2():
    html = ""
    html += af_htmlformsubmitbutton()
    html += "</form>"
    return html

def af_redirect(link="flangehandle.flangehandle"):
    return redirect(url_for(link))

def af_htmlhandletemplate(linkdownload="/flangehandle/download", linkupload="/flangehandle/upload", linkimages="/flangehandle/images", name="flange.mdb", linkback="/flangehandle"):
    html = ""
    html += """
        <!doctype html>
        <html>

            <head>
                <title>Upload/Download File</title>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <link href="{{ url_for('static', filename='bootstrap.min.css') }}" rel="stylesheet" integrity="sha384-sRIl4kxILFvY47J16cr9ZwB07vP4J8+LH7qKQnuqkuIAvNWLzeN8tE5YBujZqJLB" crossorigin="anonymous">
                <style>
                    body {
                        background-color: aliceblue;
                        background-image: url("{{ url_for('static', filename='images/flange.png') }}");
                        background-repeat: no-repeat;
                    }
                    .selectorDIV {
                        max-width: 800px;
                        margin: auto;
                        background-color: rgb(249, 252, 255);
                        margin-top: 20px;
                        padding: 20px;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                        text-align: center;
                    }
                    .chooseDIV div {
                        background-color: #fff;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                        width: 50%;
                        text-align: center;
                        display: flex;
                        flex-direction: column;
                        justify-content: center;
                        align-items: center;
                        padding:10px;
                        margin:5px;
                    }
                    .chooseDIV div form input {
                        margin-bottom: 10px;
                    }
                </style>
            </head>

            <body>
                <div class="selectorDIV" style="
                        font-size: 40px;
                        padding: 5px;
                        margin: 10px;
                        margin-left: 20px;
                        position: absolute;
                    ">
                    <a href=\""""+linkback+"""\" style="
                        text-decoration: none;
                    " >🔙</a>
                </div>
                <div class="container selectorDIV">
                    <div class="d-flex justify-content-center chooseDIV">
                        <div>
                            <h1>📩Download File</h1>
                            <div class="form-control">
                                <a href=\""""+linkdownload+"""\">Download """+name+"""</a>
                            </div>
                        </div>
                        <div>
                            <h1>⬆️Upload File</h1>
                            <form action=\""""+linkupload+"""\" method="post" enctype="multipart/form-data">
                                <input type="file" name="file" class="form-control">
                                <input type="submit" value="Upload" class="btn btn-warning">
                            </form>
                        </div>
                    </div>"""

    html += af_render_flashed_messages()
    
    if linkimages != "":
        html += """
                <div class="d-flex justify-content-center align-items-center my-4">
                    <a href=\""""+linkimages+"""\">
                        <div class="form-control text-center">
                            Images
                        </div>
                    </a>
                </div>
                """
    html += """
            </body>

        </html>
    """
    return render_template_string(html)

def af_render_flashed_messages():
    # Assume get_flashed_messages is defined and returns a list of messages
    messages = get_flashed_messages()
    if not messages:
        return ""

    # Start building the HTML string
    html = "<div class='flashed-messages'><br/><ul>"
    
    # Add each message as a list item
    for message in messages:
        html += f"<li>{message}</li>"
    
    # Close the unordered list and the div
    html += "</ul></div>"
    
    return html


def af_htmlbulatcolor(color="#fabcde"):
    html = f"<div style='border-radius:24px;width:30px;height:30px;background-color:{color};'></div> "
    return html

def af_image_exists(image_path="static/images/employees/afwan.png"):
    # Check if the file exists and is a file (not a directory)
    return os.path.isfile(image_path)

def af_htmlimagehandletemplate(image_dir="static/images/flanges", linkrename="/flangehandle/images/rename", linkadd="/flangehandle/images/add"):

    # Path to the directory containing images
    # image_dir = os.path.join('static', 'images', 'flanges')

    # List all image files in the directory
    images = [
        {'name': file_name, 'src': f'/{image_dir}/{file_name}'}
        for file_name in os.listdir(image_dir)
        if os.path.isfile(os.path.join(image_dir, file_name))
    ]

    html = """
        <!doctype html>
        <html>

        <head>
            <title>Upload/Download File</title>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <link href="{{ url_for('static', filename='bootstrap.min.css') }}" rel="stylesheet"
                integrity="sha384-sRIl4kxILFvY47J16cr9ZwB07vP4J8+LH7qKQnuqkuIAvNWLzeN8tE5YBujZqJLB" crossorigin="anonymous">
            <style>
                body {
                    background-color: aliceblue;
                    background-image: url("{{ url_for('static', filename='images/flange.png') }}");
                    background-repeat: no-repeat;
                }

                .selectorDIV {
                    max-width: 800px;
                    margin: auto;
                    background-color: rgb(249, 252, 255);
                    margin-top: 20px;
                    padding: 20px;
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                    text-align: center;
                }

                .chooseDIV div {
                    background-color: #fff;
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                    width: 50%;
                    text-align: center;
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                    padding: 10px;
                    margin: 5px;
                }

                .chooseDIV div form input {
                    margin-bottom: 10px;
                }

                .title {
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                }

                .title h1 {
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                    background-color: rgba(255, 255, 255, 0.9);
                    /* border-radius: 12px; */
                    width: fit-content;
                    padding: 0.5rem;
                }
            </style>
        </head>

        <body class="container my-3" style="position: relative;">
            <div class="title my-4">
                <h1 class="text-center">
                    Images
                </h1>
            </div>
            <a style="position: absolute;
            right: 10px;
            top: 20px;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            border-radius: 12px;
            background-color: lightblue;
            font-weight: bold;
            padding: 10px;
            color: black;
            text-decoration: none;" href=\""""+linkadd+"""\">
                <div>
                    + Add image
                </div>
            </a>

            <div class="row">
                {% for im in images %}
                <div class="col-sm-6 col-md-4 col-lg-3 mb-4">
                    <div class="card"><img src="{{ im.src }}" class="card-img-top" alt="{{ im.name }}">
                        <div class="card-body text-center">
                            <h5 class="card-title">{{ im.name }}</h5><!-- Form for renaming the image -->
                            <form action=\""""+linkrename+"""\" method="post"><input type="hidden" name="old_name"
                                    value="{{ im.name }}"><input type="text" name="new_name" placeholder="New name"><button
                                    type="submit">Rename</button></form>
                        </div>
                    </div>
                </div>
                {% endfor %}
            </div>
        </body>

        </html>
    """

    return render_template_string(html, images=images)

def af_htmlimagehandletemplateadd(endpointupload="/flangehandle/images/upload", endpointback="/flangehandle/images"):
    html = """
<!doctype html>
<html>

<head>
    <title>Upload/Download File</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="{{ url_for('static', filename='bootstrap.min.css') }}" rel="stylesheet"
        integrity="sha384-sRIl4kxILFvY47J16cr9ZwB07vP4J8+LH7qKQnuqkuIAvNWLzeN8tE5YBujZqJLB" crossorigin="anonymous">
    <style>
        body {
            background-color: aliceblue;
            background-image: url("{{ url_for('static', filename='images/flange.png') }}");
            background-repeat: no-repeat;
        }

        .selectorDIV {
            max-width: 800px;
            margin: auto;
            background-color: rgb(249, 252, 255);
            margin-top: 20px;
            padding: 20px;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            text-align: center;
        }

        .chooseDIV div {
            background-color: #fff;
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            width: 50%;
            text-align: center;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            padding: 10px;
            margin: 5px;
        }

        .chooseDIV div form input {
            margin-bottom: 10px;
        }

        .title {
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
        }

        .title h1 {
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
            background-color: rgba(255, 255, 255, 0.9);
            /* border-radius: 12px; */
            width: fit-content;
            padding: 0.5rem;
        }
    </style>
</head>

<body class="container my-3" style="position: relative;">
    <div class="title my-4">
        <h1 class="text-center">
            Add Images
        </h1>
    </div>
    <a style="position: absolute;
    right: 10px;
    top: 20px;
    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
    border-radius: 12px;
    background-color: lightblue;
    font-weight: bold;
    padding: 10px;
    color: black;
    text-decoration: none;" href=\""""+endpointback+"""\">
        <div>
            Back
        </div>
    </a>

    <form style="max-width: 500px;" class="container w-50" action=\""""+endpointupload+"""\" method="post" enctype="multipart/form-data">
        <div class="row justify-content-center align-items-center my-4">
            <div class="w-50 mb-2">
                <input type="text" name="codename" class="form-control" placeholder="codename"
                    style="text-transform:uppercase;" required />
            </div>
            <div class="w-75 mb-2">
                <input type="file" name="image" accept="image/*" class="form-control" required />
            </div>
            <!-- <div class="d-flex justify-content-center align-items-center mb-2">
                <button class="btn btn-primary">
                    Upload
                </button>
            </div> -->
        </div>
        <div class="my-4 text-center">
            <button type="submit" class="form-control btn btn-warning w-50" value="Submit">
                Submit
            </button>
        </div>
    </form>
</body>

</html>
"""
    return render_template_string(html)

def af_htmlimagehandletemplateupload(image_dir="/static/images/flanges", allowed_extensions={'jpeg','png','jpg'}, redirectlink="/flangehandle/images"):
    # Directory to save uploaded images
    # UPLOAD_FOLDER = os.path.join(os.getcwd(), 'static', 'images', 'flanges')
    # UPLOAD_FOLDER = image_dir
    UPLOAD_FOLDER = os.path.join(os.getcwd()) + image_dir
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    # Allowed file extensions
    # ALLOWED_EXTENSIONS = {'jpeg','png','jpg'}
    ALLOWED_EXTENSIONS = allowed_extensions

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    if request.method == 'POST':
        codename = request.form['codename'].upper()
        file = request.files['image']

        if file and allowed_file(file.filename):
            # Use secure_filename to safely handle user input filenames
            filename = secure_filename(f"{codename}.jpeg")
            filepath = os.path.join(UPLOAD_FOLDER, filename)

            # Save the uploaded file
            file.save(filepath)

            # (Optional) Convert to JPEG in case it's not, use PIL
            with Image.open(filepath) as img:
                # if img.format != 'JPEG':
                #     img = img.convert('RGB')  # Convert to RGB mode if needed
                #     img.save(filepath, 'JPEG')
                
                img = img.convert('RGB')  # Convert to RGB mode if needed
                img.save(filepath, 'JPEG')

            # return redirect(url_for('upload_file'))

    return redirect(redirectlink)

def af_htmlimagehandletemplaterename(image_dir="static/images/flanges", redirectlink="/flangehandle/images"):
    old_name = request.form['old_name']
    new_name = request.form['new_name']

    # Ensure the old file exists and new file doesn't
    old_path = image_dir + "/" + old_name
    new_path = image_dir + "/" + new_name


    if os.path.exists(old_path) and not os.path.exists(new_path):
        os.rename(old_path, new_path)
        # Update the image list accordingly if needed
    else:
        # Handle error (e.g., file not found or new file name already used)
        pass

    return redirect(redirectlink)